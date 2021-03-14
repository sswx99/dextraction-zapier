import scraper
import openpyxl

TOTAL_ROW = 1211
APP_NAME_COLUMN = 4
TEMPLATE_COLUMN_START = 5

def write_zaps():
    workbook = openpyxl.load_workbook("data/ZapierApps.xlsx")
    sheet = workbook.active

    for row in range(3, TOTAL_ROW):
        app = sheet.cell(row=row, column=APP_NAME_COLUMN).value
        slug = scraper.get_zap_slug(app)
        templates = scraper.get_zap_templates(slug)
        
        column = 0
        for template in templates:
          sheet.cell(row, TEMPLATE_COLUMN_START + column, value=template)
    
    workbook.save("data/ZapierAppsWithTemplates.xlsx")



if __name__ == "__main__":
  write_zaps()